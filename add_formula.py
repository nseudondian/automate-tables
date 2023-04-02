from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

wb = load_workbook('pivot_table.xlsx')
sheet = wb['Report']

min_col = wb.active.min_column
max_col = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

# print(min_col)
# print(max_col)
# print(min_row)
# print(max_row)

barchart = BarChart()
data = Reference(sheet, min_col=min_col+1, max_col=max_col, min_row=min_row, max_row = max_row)

categories = Reference(sheet, min_col=min_col, max_col=min_col, min_row=min_row+1, max_row = max_row)

barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)

sheet.add_chart(barchart, "B12")

barchart.title = "Sales by Product line"
barchart.style = 3


# Create a formula
# sheet['B8'] = '=SUM(B6:B7)'
# sheet['B8'].style = 'Currency'

for i in range(min_col + 1, max_col+1):
    letter = get_column_letter(i)
    sheet[f'{letter}{max_row+1}'].style = 'Currency'
    sheet[f'{letter}{max_row+1}'] = f'=SUM({letter}{min_row+1}:{letter}{max_row})'
    



wb.save('report.xlsx')