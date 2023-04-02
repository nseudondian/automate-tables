from openpyxl import load_workbook
# from openpyxl.styles import Font
# import os
# import sys

# application_path = os.path.dirname(sys.executable)
# month = input("Enter month to consider: ")

# input_path = os.path.join(application_path, 'pvot_table.xlsx')

# wb = load_workbook(input_path)
# sheet = wb['Report']

# # sheet['A1'] = 'Sales Report'
# # sheet["A2"] = month

# # sheet['A1'].font = Font('Calibri', bold=True, size=20 )
# # sheet['A2'].font = Font('Calibri', bold=True, size=10 )

# min_col = wb.active.min_column
# max_col = wb.active.max_column
# min_row = wb.active.min_row
# max_row = wb.active.max_row

# barchart = BarChart()
# data = Reference(sheet, min_col=min_col+1, max_col=max_col, min_row=min_row, max_row = max_row)

# categories = Reference(sheet, min_col=min_col, max_col=min_col, min_row=min_row+1, max_row = max_row)

# barchart.add_data(data, titles_from_data=True)
# barchart.set_categories(categories)

# sheet.add_chart(barchart, "B12")

# barchart.title = "Sales by Product line"
# barchart.style = 3

# sheet['A1'] = 'Sales Report'
# sheet['A2'] = month
# sheet['A1'].font = Font('Arial', bold=True, size=20)
# sheet['A2'].font = Font('Arial', bold=True, size=10)


# output_path = os.path.join(application_path,f'report_{month}.xlsx' )

# wb.save(output_path)